import openpyxl
import pandas as pd
import json
import time
import os
import win32com.client as win32

file_path = os.path.abspath('master.xlsx') 

if not os.path.exists(file_path):
    raise FileNotFoundError(f"Excel file not found: {file_path}")

workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

new_interest_y = "3,00%"  
new_number_of_periods = 500  
new_principal = 260000  

sheet.cell(row=3, column=5).value = new_interest_y
sheet.cell(row=3, column=7).value = new_number_of_periods
sheet.cell(row=3, column=8).value = new_principal

workbook.save(file_path)

excel = win32.gencache.EnsureDispatch('Excel.Application')
workbook = excel.Workbooks.Open(file_path)
excel.Visible = False  

workbook.Save()  
workbook.Close(SaveChanges=True) 

excel.Quit()

time.sleep(1)

df = pd.read_excel(file_path)

csv_file_path = 'updated_data.csv'
df.to_csv(csv_file_path, index=False)

df_updated = pd.read_csv(csv_file_path, skiprows=1)

print("Columns in the updated DataFrame:", df_updated.columns)
print("Sample data from the updated DataFrame:\n", df_updated.head())

payment_column = 'payment'

if df_updated[payment_column].dropna().empty:
    raise ValueError("The 'payment' column contains only NaN values")

payment_value = df_updated[payment_column].dropna().values[0]

payment_value = round(abs(payment_value), 2)

output_data = {
    "payment": payment_value,
    "interest_y": new_interest_y,
    "number_of_periods": new_number_of_periods,
    "principal": new_principal
}

json_output = json.dumps(output_data, indent=4)

print(json_output)
