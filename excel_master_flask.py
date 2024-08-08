from flask import Flask, request, jsonify
import openpyxl
import pandas as pd
import json
import time
import os
import shutil
import pythoncom
import win32com.client as win32
from collections import OrderedDict
from datetime import datetime

app = Flask(__name__)

@app.route('/excel_master/update_excel', methods=['POST'])
def update_excel():
    pythoncom.CoInitialize()

    try:
        data = request.get_json()
        new_interest_y = data["interest_y"]
        new_number_of_periods = data["number_of_periods"]
        new_principal = data["principal"]

        file_path = os.path.abspath('master.xlsx')

        if not os.path.exists(file_path):
            return jsonify({"error": "Excel file not found"}), 404

        backup_folder = os.path.abspath('backups')
        if not os.path.exists(backup_folder):
            os.makedirs(backup_folder)

        backup_file_path = os.path.join(backup_folder, f'master_backup_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx')
        shutil.copyfile(file_path, backup_file_path)

        backups = sorted(os.listdir(backup_folder))
        if len(backups) > 10:
            oldest_backup = os.path.join(backup_folder, backups[0])
            os.remove(oldest_backup)

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        sheet.cell(row=3, column=5).value = new_interest_y
        sheet.cell(row=3, column=7).value = new_number_of_periods
        sheet.cell(row=3, column=8).value = new_principal

        workbook.save(file_path)

        excel = win32.gencache.EnsureDispatch('Excel.Application')
        workbook = excel.Workbooks.Open(file_path)
        excel.Visible = False

        workbook.Save()
        workbook.Close(SaveChanges=True)
        excel.Quit()

        time.sleep(1)

        df = pd.read_excel(file_path)
        csv_file_path = 'updated_data.csv'
        df.to_csv(csv_file_path, index=False)

        df_updated = pd.read_csv(csv_file_path, skiprows=1)

        payment_column = 'payment'

        if df_updated[payment_column].dropna().empty:
            return jsonify({"error": "The 'payment' column contains only NaN values"}), 400

        payment_value = df_updated[payment_column].dropna().values[0]
        payment_value = round(abs(payment_value), 2)

        output_data = OrderedDict([
            ("payment", payment_value),
            ("interest_y", new_interest_y),
            ("number_of_periods", new_number_of_periods),
            ("principal", new_principal)
        ])

        return jsonify(output_data)

    finally:
        pythoncom.CoUninitialize()

if __name__ == '__main__':
    app.run(debug=True)
